from werkzeug.wrappers import Request, Response
from werkzeug.serving import run_simple
from jsonrpcserver import methods
from xlrd.biffh import XLRDError
from datetime import datetime
from openpyxl.styles import PatternFill, Border
import pandas as pd
import os,sys

file_path = ""
selected_sheet = ""
selected_column = ""
selected_row = None
selected_data = ""
data = {}

@methods.add
def read_sheetname(thepath):
	global file_path,data,selected_data
	if os.path.isdir(thepath):
		return {"error":"非文件！:{}".format(file_path)};
	file_path = thepath
	file = os.path.basename(file_path).split(".")
	try:
		data = pd.ExcelFile(file_path,na_filter=False)# as excelfile: "{}{}{}".format(path,os.sep,file)
	except FileNotFoundError:
		return {"error":"無此文件！:{}".format(file_path)};
	except XLRDError:
		return {"error":"非Excel文件！:{}".format(file_path)};
	return {"data":data.sheet_names,"filename":file[0]}

@methods.add
def read_column():
	global selected_sheet,selected_data,data
	selected_data = data.parse(selected_sheet,na_filter=False)
	return list(selected_data.columns)

@methods.add
def read_row():
	global selected_data,selected_column
	return {"data":list(selected_data[selected_column[0]]),"row_name":selected_column[0]}
	
@methods.add
def update_sheetname(thesheetname):
	global selected_sheet
	selected_sheet = thesheetname
	
@methods.add
def update_column(*thecolumn):
	global selected_column
	selected_column = thecolumn
	selected_data[selected_column[0]] = selected_data[selected_column[0]].str.upper()
	
@methods.add
def update_row(*therow):
	global selected_row
	selected_row = therow

@methods.add
def write_excel():
	global file_path,selected_column,selected_row,selected_data
	file = os.path.basename(file_path).split(".")
	outputfilename = "{}{}{}_{}.{}".format(os.path.dirname(file_path),os.sep,file[0],int(datetime.utcnow().timestamp()),file[1])
	newdata = pd.DataFrame()
	for column in selected_column:
		newdata[column] = selected_data[column]
	newdata.set_index(selected_column[0])
	if selected_row !=None:
		newdata = newdata[newdata[selected_column[0]].isin(selected_row)]
		newdata.insert(0,"序號", range(1,newdata.shape[0]+1))
	with pd.ExcelWriter(outputfilename) as writer:
		newdata.to_excel(writer, sheet_name='Sheet1',index=False)
		fill_color = PatternFill(fgColor = 'D0D4F0',fill_type='solid')
		target = writer.sheets['Sheet1']
		target_max_column = len(selected_column)
		target_max_row = newdata.shape[0]
		for row_index in range(1,target_max_row+1):
			if row_index%2 != 0:
				for column_index in range(1,target_max_column+1):
					target.cell(column=column_index,row=row_index).fill = fill_color
		#cell_format = writer.Workbook.add_format({'type': '3_color_scale'});
		#test.set_row(1, None, cell_format)
	return "文件處理完畢，文件名:{}".format(outputfilename)
	
@Request.application
def application(request):
	r = methods.dispatch(request.data.decode())
	return Response(str(r), r.http_status, mimetype='application/json')

if __name__ == '__main__':
    run_simple('localhost', 7749, application)